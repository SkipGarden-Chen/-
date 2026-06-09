from __future__ import annotations

import argparse
import csv
import math
from pathlib import Path
from xml.sax.saxutils import escape


CANVAS_WIDTH = 1420
CANVAS_HEIGHT = 860
MARGIN_LEFT = 92
MARGIN_RIGHT = 290
MARGIN_TOP = 84
MARGIN_BOTTOM = 128
PLOT_WIDTH = CANVAS_WIDTH - MARGIN_LEFT - MARGIN_RIGHT
PLOT_HEIGHT = CANVAS_HEIGHT - MARGIN_TOP - MARGIN_BOTTOM


LITHOLOGY_STYLES = {
    "sandstone": {"label": "砂岩", "fill": "#f4d27a", "pattern": "sandstone"},
    "mudstone": {"label": "泥岩/页岩", "fill": "#6f7787", "pattern": "mudstone"},
    "limestone": {"label": "灰岩", "fill": "#c9d6df", "pattern": "limestone"},
    "conglomerate": {"label": "砾岩", "fill": "#d6b48a", "pattern": "conglomerate"},
    "granite": {"label": "花岗岩", "fill": "#efb5b5", "pattern": "granite"},
    "basalt": {"label": "玄武岩", "fill": "#4f5965", "pattern": "basalt"},
    "default": {"label": "其他岩性", "fill": "#d8dee6", "pattern": "default"},
}


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def read_xlsx_sheet(path: Path, sheet_name: str) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 需要 openpyxl，请改用 CSV 表格目录。") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record: dict[str, str] = {}
        empty = True
        for header, value in zip(headers, row):
            if not header:
                continue
            text = "" if value is None else str(value).strip()
            if text:
                empty = False
            record[header] = text
        if not empty:
            records.append(record)
    workbook.close()
    return records


class TableSource:
    def __init__(self, path: Path):
        self.path = path
        self.is_excel = path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}

    def read(self, table_name: str) -> list[dict[str, str]]:
        if self.is_excel:
            return read_xlsx_sheet(self.path, table_name)
        return read_csv(self.path / f"{table_name}.csv")


def to_float(value: str | None, default: float | None = None) -> float | None:
    if value is None:
        return default
    value = str(value).strip()
    if not value:
        return default
    try:
        return float(value)
    except ValueError:
        return default


def clean(value: str | None, default: str = "") -> str:
    if value is None:
        return default
    value = str(value).strip()
    return value if value else default


def lithology_key(name: str) -> str:
    text = name.lower()
    if "砂" in text or "sand" in text:
        return "sandstone"
    if "泥" in text or "页" in text or "shale" in text or "mud" in text:
        return "mudstone"
    if "灰岩" in text or "白云岩" in text or "limestone" in text:
        return "limestone"
    if "砾" in text or "conglomerate" in text:
        return "conglomerate"
    if "花岗" in text or "granite" in text:
        return "granite"
    if "玄武" in text or "basalt" in text:
        return "basalt"
    return "default"


def nice_tick_interval(length: float) -> float:
    if length <= 50:
        return 10
    if length <= 120:
        return 20
    if length <= 250:
        return 50
    return 100


def lerp(a: float, b: float, t: float) -> float:
    return a + (b - a) * t


def safe_id(value: str) -> str:
    cleaned = []
    for char in value:
        if char.isalnum() or char in {"_", "-"}:
            cleaned.append(char)
        else:
            cleaned.append("_")
    return "".join(cleaned) or "item"


class Topography:
    def __init__(self, points: list[tuple[float, float]], total_length: float):
        if not points:
            points = [(0.0, 0.0), (total_length, 0.0)]
        points = sorted(points)
        if points[0][0] > 0:
            points.insert(0, (0.0, points[0][1]))
        if points[-1][0] < total_length:
            points.append((total_length, points[-1][1]))
        self.points = points
        self.total_length = total_length
        self.slopes = self._build_slopes()

    def _build_slopes(self) -> list[float]:
        slopes: list[float] = []
        pts = self.points
        for index, (x, y) in enumerate(pts):
            if len(pts) == 1:
                slopes.append(0.0)
            elif index == 0:
                x1, y1 = pts[1]
                slopes.append((y1 - y) / (x1 - x) if x1 != x else 0.0)
            elif index == len(pts) - 1:
                x0, y0 = pts[index - 1]
                slopes.append((y - y0) / (x - x0) if x != x0 else 0.0)
            else:
                x0, y0 = pts[index - 1]
                x1, y1 = pts[index + 1]
                slopes.append((y1 - y0) / (x1 - x0) if x1 != x0 else 0.0)
        return slopes

    def elevation_at(self, distance: float) -> float:
        distance = max(0.0, min(self.total_length, distance))
        pts = self.points
        if distance <= pts[0][0]:
            return pts[0][1]
        for index, ((x0, y0), (x1, y1)) in enumerate(zip(pts, pts[1:])):
            if x0 <= distance <= x1:
                if x1 == x0:
                    return y0
                dx = x1 - x0
                t = (distance - x0) / dx
                m0 = self.slopes[index]
                m1 = self.slopes[index + 1]
                h00 = 2 * t**3 - 3 * t**2 + 1
                h10 = t**3 - 2 * t**2 + t
                h01 = -2 * t**3 + 3 * t**2
                h11 = t**3 - t**2
                return h00 * y0 + h10 * dx * m0 + h01 * y1 + h11 * dx * m1
        return pts[-1][1]

    def between(self, start: float, end: float) -> list[tuple[float, float]]:
        start = max(0.0, min(self.total_length, start))
        end = max(0.0, min(self.total_length, end))
        if end < start:
            start, end = end, start
        result = [(start, self.elevation_at(start))]
        for distance, elevation in self.points:
            if start < distance < end:
                result.append((distance, elevation))
        result.append((end, self.elevation_at(end)))
        return result

    def smooth_points(self, start: float, end: float, step: float | None = None) -> list[tuple[float, float]]:
        start = max(0.0, min(self.total_length, start))
        end = max(0.0, min(self.total_length, end))
        if end < start:
            start, end = end, start
        span = max(0.001, end - start)
        step = step or max(0.8, min(3.0, self.total_length / 180))
        count = max(2, int(math.ceil(span / step)) + 1)
        return [
            (lerp(start, end, index / (count - 1)), self.elevation_at(lerp(start, end, index / (count - 1))))
            for index in range(count)
        ]


class SvgBuilder:
    def __init__(
        self,
        section: dict[str, str],
        topo: Topography,
        units: list[dict[str, str]],
        structures: list[dict[str, str]],
    ):
        self.section = section
        self.topo = topo
        self.units = units
        self.structures = structures
        self.total_length = to_float(section.get("total_length_m"), topo.total_length) or topo.total_length

        elevations = [e for _, e in topo.points]
        self.elev_max = max(elevations) + max(3.0, (max(elevations) - min(elevations)) * 0.15)
        topo_range = max(1.0, max(elevations) - min(elevations))
        section_depth = max(28.0, topo_range * 1.5)
        self.elev_min = min(elevations) - section_depth

        self.x_scale = PLOT_WIDTH / self.total_length
        fit_y_scale = PLOT_HEIGHT / (self.elev_max - self.elev_min)
        self.y_scale = min(self.x_scale, fit_y_scale)
        self.y_offset = MARGIN_TOP + (PLOT_HEIGHT - (self.elev_max - self.elev_min) * self.y_scale) / 2
        self.lines: list[str] = []
        self.used_keys: set[str] = set()

    def x(self, distance: float) -> float:
        return MARGIN_LEFT + distance * self.x_scale

    def y(self, elevation: float) -> float:
        return self.y_offset + (self.elev_max - elevation) * self.y_scale

    def topo_y(self, distance: float) -> float:
        return self.y(self.topo.elevation_at(distance))

    def add(self, line: str) -> None:
        self.lines.append(line)

    def curve_path(self, points: list[tuple[float, float]]) -> str:
        screen_points = [(self.x(distance), self.y(elevation)) for distance, elevation in points]
        return self.curve_path_from_screen(screen_points)

    def curve_path_from_screen(self, points: list[tuple[float, float]]) -> str:
        if not points:
            return ""
        if len(points) == 1:
            x, y = points[0]
            return f"M {x:.2f} {y:.2f}"
        commands = [f"M {points[0][0]:.2f} {points[0][1]:.2f}"]
        for index in range(len(points) - 1):
            p0 = points[index - 1] if index > 0 else points[index]
            p1 = points[index]
            p2 = points[index + 1]
            p3 = points[index + 2] if index + 2 < len(points) else p2
            c1 = (p1[0] + (p2[0] - p0[0]) / 6, p1[1] + (p2[1] - p0[1]) / 6)
            c2 = (p2[0] - (p3[0] - p1[0]) / 6, p2[1] - (p3[1] - p1[1]) / 6)
            commands.append(
                f"C {c1[0]:.2f} {c1[1]:.2f}, {c2[0]:.2f} {c2[1]:.2f}, {p2[0]:.2f} {p2[1]:.2f}"
            )
        return " ".join(commands)

    def unit_path(self, unit: dict[str, str], start: float, end: float) -> str:
        topo_pts = self.topo.smooth_points(start, end)
        top_path = self.curve_path(topo_pts)

        dip_angle = to_float(unit.get("dip_angle_deg"), 0.0) or 0.0
        rendered_dip = self.bedding_dip_for_render(to_float(unit.get("dip_direction_deg"), None), dip_angle)
        rendered_dip = max(-82.0, min(82.0, rendered_dip))
        slope = -math.tan(math.radians(rendered_dip))
        direction = self.down_dip_direction(slope)

        start_tail = self.line_endpoint_to_plot(start, self.topo.elevation_at(start), slope, direction)
        end_tail = self.line_endpoint_to_plot(end, self.topo.elevation_at(end), slope, direction)
        return (
            f"{top_path} "
            f"L {self.x(end_tail[0]):.2f} {self.y(end_tail[1]):.2f} "
            f"L {self.x(start_tail[0]):.2f} {self.y(start_tail[1]):.2f} Z"
        )

    def render(self) -> str:
        self.add_header()
        self.add_defs()
        self.add_background()
        self.add_grid_and_axes()
        self.add_units()
        self.add_topography()
        self.add_structures()
        self.add_unit_labels_and_attitudes()
        self.add_legend()
        self.add_footer()
        return "\n".join(self.lines)

    def add_header(self) -> None:
        self.add(
            f'<svg xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" width="{CANVAS_WIDTH}" '
            f'height="{CANVAS_HEIGHT}" viewBox="0 0 {CANVAS_WIDTH} {CANVAS_HEIGHT}">'
        )

    def add_defs(self) -> None:
        self.add("<defs>")
        self.add(
            '<filter id="shadow" x="-20%" y="-20%" width="140%" height="140%">'
            '<feDropShadow dx="0" dy="2" stdDeviation="3" flood-color="#1f2933" flood-opacity="0.12"/>'
            "</filter>"
        )
        for key, style in LITHOLOGY_STYLES.items():
            self.add_lithology_pattern(f"pat_{key}", key)
        self.add("</defs>")

    def add_lithology_pattern(self, pattern_id: str, key: str, angle: float | None = None) -> None:
        style = LITHOLOGY_STYLES.get(key, LITHOLOGY_STYLES["default"])
        transform = "" if angle is None else f' patternTransform="rotate({angle:.2f})"'
        self.add(f'<pattern id="{pattern_id}" width="18" height="18" patternUnits="userSpaceOnUse"{transform}>')
        self.add(f'<rect width="18" height="18" fill="{style["fill"]}"/>')
        if key == "sandstone":
            self.add('<circle cx="5" cy="5" r="1.1" fill="#8a6d3b" opacity="0.75"/>')
            self.add('<circle cx="14" cy="12" r="1" fill="#8a6d3b" opacity="0.65"/>')
        elif key == "mudstone":
            self.add('<path d="M0 5 H18 M0 11 H18 M0 17 H18" stroke="#3f4854" stroke-width="0.9" opacity="0.72"/>')
        elif key == "limestone":
            self.add('<path d="M0 6 H18 M0 12 H18 M9 0 V6 M4 6 V12 M13 12 V18" stroke="#6d8291" stroke-width="0.9" opacity="0.75"/>')
        elif key == "conglomerate":
            self.add('<circle cx="5" cy="6" r="3" fill="none" stroke="#8b6f47" stroke-width="1"/>')
            self.add('<circle cx="14" cy="13" r="2.5" fill="none" stroke="#8b6f47" stroke-width="1"/>')
        elif key == "granite":
            self.add('<path d="M3 4 L6 7 M12 3 L15 6 M4 14 L8 11 M12 13 L16 16" stroke="#8e4d5a" stroke-width="1.1"/>')
            self.add('<circle cx="9" cy="9" r="1" fill="#8e4d5a"/>')
        elif key == "basalt":
            self.add('<path d="M0 0 L18 18 M-4 8 L8 20 M10 -2 L22 10" stroke="#25303a" stroke-width="1" opacity="0.65"/>')
        else:
            self.add('<path d="M0 0 L18 18 M18 0 L0 18" stroke="#9aa5b1" stroke-width="0.8" opacity="0.45"/>')
        self.add("</pattern>")

    def add_background(self) -> None:
        self.add('<rect width="100%" height="100%" fill="#f6f8fa"/>')
        self.add(
            f'<rect x="28" y="28" width="{CANVAS_WIDTH - 56}" height="{CANVAS_HEIGHT - 56}" '
            'rx="8" fill="#ffffff" stroke="#dde3ea" filter="url(#shadow)"/>'
        )
        title = clean(self.section.get("section_name"), "AI辅助地质剖面图")
        scale = clean(self.section.get("scale"), "1:1000")
        start = clean(self.section.get("start_point"), "起点")
        end = clean(self.section.get("end_point"), "终点")
        azimuth = clean(self.section.get("azimuth_deg"), "")
        subtitle = f"{start} - {end}    比例尺：{scale}"
        if azimuth:
            subtitle += f"    剖面方位：{azimuth}°"
        self.text(CANVAS_WIDTH / 2, 45, title, size=24, weight="700", anchor="middle")
        self.text(CANVAS_WIDTH / 2, 73, subtitle, size=14, fill="#52616f", anchor="middle")

    def add_grid_and_axes(self) -> None:
        left = MARGIN_LEFT
        right = MARGIN_LEFT + PLOT_WIDTH
        top = MARGIN_TOP
        bottom = MARGIN_TOP + PLOT_HEIGHT
        self.add(f'<rect x="{left}" y="{top}" width="{PLOT_WIDTH}" height="{PLOT_HEIGHT}" fill="#fbfcfd" stroke="#c7d0d9"/>')

        tick = nice_tick_interval(self.total_length)
        distance = 0.0
        while distance <= self.total_length + 0.001:
            x = self.x(distance)
            self.add(f'<line x1="{x:.2f}" y1="{top}" x2="{x:.2f}" y2="{bottom}" stroke="#e6ebf0" stroke-width="1"/>')
            self.add(f'<line x1="{x:.2f}" y1="{bottom}" x2="{x:.2f}" y2="{bottom + 6}" stroke="#52616f" stroke-width="1"/>')
            self.text(x, bottom + 23, f"{distance:g}", size=11, fill="#52616f", anchor="middle")
            distance += tick
        self.text((left + right) / 2, bottom + 48, "沿剖面距离 / m", size=13, fill="#334e68", anchor="middle")

        elev_step = self.choose_elev_step()
        first = math.floor(self.elev_min / elev_step) * elev_step
        elev = first
        while elev <= self.elev_max + 0.001:
            y = self.y(elev)
            if top <= y <= bottom:
                self.add(f'<line x1="{left}" y1="{y:.2f}" x2="{right}" y2="{y:.2f}" stroke="#edf1f5" stroke-width="1"/>')
                self.add(f'<line x1="{left - 6}" y1="{y:.2f}" x2="{left}" y2="{y:.2f}" stroke="#52616f" stroke-width="1"/>')
                self.text(left - 12, y + 4, f"{elev:g}", size=11, fill="#52616f", anchor="end")
            elev += elev_step
        self.text(left - 58, top + 18, "高程 / m", size=13, fill="#334e68", anchor="middle")

    def choose_elev_step(self) -> float:
        span = self.elev_max - self.elev_min
        if span <= 30:
            return 5
        if span <= 80:
            return 10
        if span <= 160:
            return 20
        return 50

    def add_units(self) -> None:
        for index, unit in enumerate(self.units):
            start = to_float(unit.get("from_m"), 0.0) or 0.0
            end = to_float(unit.get("to_m"), self.total_length) or self.total_length
            start = max(0.0, min(self.total_length, start))
            end = max(0.0, min(self.total_length, end))
            if end <= start:
                continue

            lith = clean(unit.get("lithology_final"), clean(unit.get("lithology_human"), "其他岩性"))
            key = lithology_key(lith)
            self.used_keys.add(key)

            unit_id = safe_id(clean(unit.get("unit_id"), f"unit_{index}"))
            clip_id = f"clip_{unit_id}"
            pattern_id = f"unit_pat_{unit_id}"
            pattern_angle = self.pattern_screen_angle(unit)
            path = self.unit_path(unit, start, end)
            self.add("<defs>")
            self.add_lithology_pattern(pattern_id, key, pattern_angle)
            self.add("</defs>")
            self.add(f'<clipPath id="{clip_id}"><path d="{path}"/></clipPath>')
            self.add(
                f'<path d="{path}" fill="url(#{pattern_id})" stroke="#596a7a" '
                'stroke-width="1.1" opacity="0.94"/>'
            )
            self.add_bedding_lines(unit, start, end, clip_id)

    def add_topography(self) -> None:
        path = self.curve_path(self.topo.smooth_points(0.0, self.total_length))
        self.add(f'<path d="{path}" fill="none" stroke="#263238" stroke-width="2.8"/>')
        self.add(f'<path d="{path}" fill="none" stroke="#ffffff" stroke-width="1.0" opacity="0.55"/>')
        for distance, elevation in self.topo.points:
            self.add(f'<circle cx="{self.x(distance):.2f}" cy="{self.y(elevation):.2f}" r="2.4" fill="#263238"/>')

    def add_bedding_lines(self, unit: dict[str, str], start: float, end: float, clip_id: str) -> None:
        thickness = to_float(unit.get("thickness_m"), None)
        dip_angle = to_float(unit.get("dip_angle_deg"), None)
        if thickness is None or thickness <= 0 or dip_angle is None:
            return

        rendered_dip = self.bedding_dip_for_render(to_float(unit.get("dip_direction_deg"), None), dip_angle)
        rendered_dip = max(-82.0, min(82.0, rendered_dip))
        slope = -math.tan(math.radians(rendered_dip))
        spacing = max(thickness, 0.2)
        start_points = self.surface_bedding_start_points(start, end, slope, spacing)
        if not start_points:
            return
        direction = self.down_dip_direction(slope)
        self.add(f'<g clip-path="url(#{clip_id})">')
        for distance, elevation in start_points:
            end_distance, end_elevation = self.line_endpoint_to_plot(distance, elevation, slope, direction)
            if abs(end_distance - distance) < 0.4:
                continue
            self.add(
                f'<line x1="{self.x(distance):.2f}" y1="{self.y(elevation):.2f}" '
                f'x2="{self.x(end_distance):.2f}" y2="{self.y(end_elevation):.2f}" '
                'stroke="#1f2933" stroke-width="1.05" opacity="0.46"/>'
            )
        self.add("</g>")

    def pattern_screen_angle(self, unit: dict[str, str]) -> float:
        dip_angle = to_float(unit.get("dip_angle_deg"), None)
        if dip_angle is None:
            return 0.0
        rendered_dip = self.bedding_dip_for_render(to_float(unit.get("dip_direction_deg"), None), dip_angle)
        rendered_dip = max(-82.0, min(82.0, rendered_dip))
        slope = -math.tan(math.radians(rendered_dip))
        return math.degrees(math.atan2(-slope * self.y_scale, self.x_scale))

    def down_dip_direction(self, slope: float) -> int:
        if slope < -1e-6:
            return 1
        if slope > 1e-6:
            return -1
        return 1

    def surface_bedding_start_points(
        self,
        start: float,
        end: float,
        slope: float,
        spacing: float,
    ) -> list[tuple[float, float]]:
        samples = self.topo.smooth_points(start, end, step=max(0.35, min(1.0, (end - start) / 120)))
        if len(samples) < 2:
            return samples

        intercepts = [(distance, elevation - slope * distance) for distance, elevation in samples]
        b_start = intercepts[0][1]
        b_end = intercepts[-1][1]
        delta_b = spacing * math.sqrt(1 + slope**2)
        if delta_b <= 0:
            return []

        direction = 1 if b_end >= b_start else -1
        targets: list[float] = [b_start]
        current = b_start + direction * delta_b
        limit = b_end
        while (direction > 0 and current <= limit + 1e-6) or (direction < 0 and current >= limit - 1e-6):
            targets.append(current)
            current += direction * delta_b

        starts: list[tuple[float, float]] = []
        last_distance = start - 1
        for target in targets:
            found = None
            for (d0, b0), (d1, b1) in zip(intercepts, intercepts[1:]):
                if d1 < last_distance:
                    continue
                if (b0 <= target <= b1) or (b1 <= target <= b0):
                    ratio = 0.0 if b1 == b0 else (target - b0) / (b1 - b0)
                    distance = lerp(d0, d1, ratio)
                    elevation = self.topo.elevation_at(distance)
                    found = (distance, elevation)
                    break
            if found and found[0] >= start - 1e-6 and found[0] <= end + 1e-6:
                if not starts or abs(found[0] - starts[-1][0]) > 0.35:
                    starts.append(found)
                    last_distance = found[0]

        if len(starts) < 2:
            horizontal_step = max(spacing, spacing / max(0.28, abs(math.cos(math.atan(slope)))))
            distance = start
            starts = []
            while distance <= end + 1e-6:
                starts.append((distance, self.topo.elevation_at(distance)))
                distance += horizontal_step
        return starts[:120]

    def line_endpoint_to_plot(
        self,
        distance: float,
        elevation: float,
        slope: float,
        direction: int,
    ) -> tuple[float, float]:
        side_limit = (self.total_length - distance) if direction > 0 else distance
        side_limit = max(0.0, side_limit)
        vertical_rate = slope * direction
        if vertical_rate < -1e-6:
            base_limit = (self.elev_min - elevation) / vertical_rate
            base_limit = max(0.0, base_limit)
        else:
            base_limit = side_limit
        travel = min(side_limit, base_limit)
        end_distance = distance + direction * travel
        end_elevation = elevation + slope * direction * travel
        return end_distance, max(self.elev_min, end_elevation)

    def add_structures(self) -> None:
        for structure in self.structures:
            stype = clean(structure.get("type")).lower()
            final_result = clean(structure.get("final_result"), clean(structure.get("type"), "构造"))
            if stype == "fault" or "断层" in final_result:
                self.draw_fault(structure, final_result)
            elif stype == "joint_zone" or "节理" in final_result:
                self.draw_joint_zone(structure, final_result)
            elif stype == "fold" or "褶皱" in final_result:
                self.draw_fold(structure, final_result)
            else:
                self.draw_generic_structure(structure, final_result)

    def draw_fault(self, structure: dict[str, str], label: str) -> None:
        distance = to_float(structure.get("distance_m"), None)
        if distance is None:
            return
        distance = max(0.0, min(self.total_length, distance))
        x0 = self.x(distance)
        y0 = self.topo_y(distance)
        bottom = self.y(self.elev_min)
        apparent = self.apparent_dip(
            to_float(structure.get("dip_direction_deg"), None),
            to_float(structure.get("dip_angle_deg"), 65.0) or 65.0,
        )
        tan_app = math.tan(math.radians(max(8.0, min(85.0, abs(apparent)))))
        sign = 1 if apparent >= 0 else -1
        horizontal_m = ((bottom - y0) / self.y_scale) / tan_app
        x1 = x0 + sign * horizontal_m * self.x_scale
        x1 = max(MARGIN_LEFT, min(MARGIN_LEFT + PLOT_WIDTH, x1))
        self.add(f'<line x1="{x0:.2f}" y1="{y0:.2f}" x2="{x1:.2f}" y2="{bottom:.2f}" stroke="#c0392b" stroke-width="3"/>')
        self.add(f'<line x1="{x0:.2f}" y1="{y0:.2f}" x2="{x1:.2f}" y2="{bottom:.2f}" stroke="#ffffff" stroke-width="1.1" stroke-dasharray="8 8"/>')
        self.text(x0 + 8, max(MARGIN_TOP + 18, y0 - 12), label, size=13, weight="700", fill="#a93226", anchor="start")

    def draw_joint_zone(self, structure: dict[str, str], label: str) -> None:
        start = to_float(structure.get("from_m"), None)
        end = to_float(structure.get("to_m"), None)
        distance = to_float(structure.get("distance_m"), None)
        if start is None or end is None:
            if distance is None:
                return
            start, end = distance - 3, distance + 3
        start = max(0.0, min(self.total_length, start))
        end = max(0.0, min(self.total_length, end))
        if end < start:
            start, end = end, start
        x = self.x(start)
        width = max(4.0, (end - start) * self.x_scale)
        y_top = min(self.topo_y(start), self.topo_y(end))
        y_bottom = self.y(self.elev_min)
        self.add(f'<rect x="{x:.2f}" y="{y_top:.2f}" width="{width:.2f}" height="{y_bottom - y_top:.2f}" fill="#6c5ce7" opacity="0.13" stroke="#6c5ce7" stroke-width="1.3" stroke-dasharray="4 4"/>')
        self.text(x + width / 2, y_top - 8, label, size=12, fill="#4b3fbf", anchor="middle")

    def draw_fold(self, structure: dict[str, str], label: str) -> None:
        distance = to_float(structure.get("distance_m"), None)
        if distance is None:
            return
        x0 = self.x(distance)
        y0 = self.topo_y(distance) + 54
        points = []
        for i in range(40):
            t = i / 39
            x = x0 - 42 + t * 84
            y = y0 + math.sin(t * math.pi * 2) * 13
            points.append(f"{x:.2f},{y:.2f}")
        self.add(f'<polyline points="{" ".join(points)}" fill="none" stroke="#8e44ad" stroke-width="2.2"/>')
        self.text(x0, y0 - 24, label, size=12, fill="#71368a", anchor="middle")

    def draw_generic_structure(self, structure: dict[str, str], label: str) -> None:
        distance = to_float(structure.get("distance_m"), None)
        if distance is None:
            return
        x = self.x(distance)
        self.add(f'<line x1="{x:.2f}" y1="{self.topo_y(distance):.2f}" x2="{x:.2f}" y2="{self.y(self.elev_min):.2f}" stroke="#2f80ed" stroke-width="1.8" stroke-dasharray="5 5"/>')
        self.text(x + 6, self.topo_y(distance) - 8, label, size=12, fill="#1f5fbf", anchor="start")

    def add_unit_labels_and_attitudes(self) -> None:
        for unit in self.units:
            start = to_float(unit.get("from_m"), 0.0) or 0.0
            end = to_float(unit.get("to_m"), self.total_length) or self.total_length
            if end <= start:
                continue
            mid = (start + end) / 2
            width_px = max(1, (end - start) * self.x_scale)
            label = clean(unit.get("unit_id"), "")
            lith = clean(unit.get("lithology_final"), clean(unit.get("lithology_human"), ""))
            line1 = f"{label}  {lith}".strip()
            structure = clean(unit.get("structure"), "")
            x = self.x(mid)
            y = min(self.topo_y(mid) + 58, self.y(self.elev_min) - 52)
            max_chars = 8 if width_px < 150 else 13
            self.multiline_text(x, y, self.wrap_label(line1, max_chars), size=13, weight="700", anchor="middle")
            if structure:
                self.multiline_text(x, y + 34, self.wrap_label(structure, max_chars), size=11, fill="#334e68", anchor="middle")
            self.draw_attitude_symbol(unit, x, min(y + 72, self.y(self.elev_min) - 24), width_px)

    def draw_attitude_symbol(self, unit: dict[str, str], x: float, y: float, unit_width: float) -> None:
        dip_angle = to_float(unit.get("dip_angle_deg"), None)
        if dip_angle is None:
            return
        dip_direction = to_float(unit.get("dip_direction_deg"), None)
        rendered_dip = self.bedding_dip_for_render(dip_direction, dip_angle)
        length = max(34, min(72, unit_width * 0.42))
        tan_app = math.tan(math.radians(max(-80, min(80, rendered_dip))))
        dy = tan_app * (length / self.x_scale) * self.y_scale
        dy = max(-34, min(34, dy))
        x1, x2 = x - length / 2, x + length / 2
        y1, y2 = y - dy / 2, y + dy / 2
        self.add(f'<line x1="{x1:.2f}" y1="{y1:.2f}" x2="{x2:.2f}" y2="{y2:.2f}" stroke="#1f2933" stroke-width="2"/>')
        self.add(f'<line x1="{x:.2f}" y1="{y - 6:.2f}" x2="{x:.2f}" y2="{y + 6:.2f}" stroke="#1f2933" stroke-width="1.4"/>')
        dip_direction_text = clean(unit.get("dip_direction_deg"), "")
        label = f"{dip_direction_text}°∠{dip_angle:g}°" if dip_direction_text else f"∠{dip_angle:g}°"
        self.text(x, y + 22, label, size=10.5, fill="#334e68", anchor="middle")

    def apparent_dip(self, dip_direction: float | None, dip_angle: float) -> float:
        section_azimuth = to_float(self.section.get("azimuth_deg"), None)
        if dip_direction is None or section_azimuth is None:
            return dip_angle
        delta = math.radians(((dip_direction - section_azimuth + 180) % 360) - 180)
        tan_app = math.tan(math.radians(dip_angle)) * math.cos(delta)
        return math.degrees(math.atan(tan_app))

    def bedding_dip_for_render(self, dip_direction: float | None, dip_angle: float) -> float:
        section_azimuth = to_float(self.section.get("azimuth_deg"), None)
        if dip_direction is None or section_azimuth is None:
            return dip_angle
        delta = math.radians(((dip_direction - section_azimuth + 180) % 360) - 180)
        sign = 1 if math.cos(delta) >= 0 else -1
        return sign * abs(dip_angle)

    def add_legend(self) -> None:
        x = MARGIN_LEFT + PLOT_WIDTH + 34
        y = MARGIN_TOP + 8
        self.text(x, y, "图例", size=17, weight="700", anchor="start")
        y += 28
        for key in sorted(self.used_keys):
            style = LITHOLOGY_STYLES.get(key, LITHOLOGY_STYLES["default"])
            self.add(f'<rect x="{x}" y="{y}" width="28" height="18" fill="url(#pat_{key})" stroke="#596a7a"/>')
            self.text(x + 38, y + 14, style["label"], size=12.5, fill="#334e68", anchor="start")
            y += 30
        y += 10
        self.add(f'<line x1="{x}" y1="{y}" x2="{x + 36}" y2="{y}" stroke="#263238" stroke-width="2.6"/>')
        self.text(x + 46, y + 4, "地形线", size=12.5, fill="#334e68", anchor="start")
        y += 26
        self.add(f'<line x1="{x}" y1="{y}" x2="{x + 36}" y2="{y}" stroke="#c0392b" stroke-width="3"/>')
        self.text(x + 46, y + 4, "断层", size=12.5, fill="#334e68", anchor="start")
        y += 26
        self.add(f'<rect x="{x}" y="{y - 9}" width="36" height="18" fill="#6c5ce7" opacity="0.13" stroke="#6c5ce7" stroke-dasharray="4 4"/>')
        self.text(x + 46, y + 4, "节理密集带", size=12.5, fill="#334e68", anchor="start")
        y += 28
        self.add(f'<line x1="{x}" y1="{y + 8}" x2="{x + 36}" y2="{y - 8}" stroke="#1f2933" stroke-width="1.2" opacity="0.55"/>')
        self.text(x + 46, y + 4, "层面线：间距=厚度，角度=产状", size=12.5, fill="#334e68", anchor="start")

        y += 54
        self.text(x, y, "说明", size=15, weight="700", anchor="start")
        notes = [
            "1. 地形线按地形点平滑插值绘制。",
            "2. 岩层内部斜线读取 thickness_m 和 dip_angle_deg。",
            "3. 岩层线和岩性花纹角度始终使用真倾角绘制。",
            "4. 图件用于比赛快速成果表达，可人工复核后修饰。",
        ]
        for note in notes:
            y += 24
            self.multiline_text(x, y, self.wrap_label(note, 17), size=11.5, fill="#52616f", anchor="start", line_height=17)

    def add_footer(self) -> None:
        bottom = MARGIN_TOP + PLOT_HEIGHT
        self.text(
            MARGIN_LEFT,
            bottom + 78,
            "数据来源：section_info.csv / topography.csv / units.csv / structures.csv",
            size=11.5,
            fill="#697889",
            anchor="start",
        )
        self.text(
            MARGIN_LEFT + PLOT_WIDTH,
            bottom + 78,
            "GeoAgent 剖面图自动生成",
            size=11.5,
            fill="#697889",
            anchor="end",
        )
        self.add("</svg>")

    def text(
        self,
        x: float,
        y: float,
        content: str,
        size: float = 12,
        fill: str = "#1f2933",
        weight: str = "400",
        anchor: str = "start",
    ) -> None:
        self.add(
            f'<text x="{x:.2f}" y="{y:.2f}" font-family="Microsoft YaHei, SimHei, Arial, sans-serif" '
            f'font-size="{size}" font-weight="{weight}" fill="{fill}" text-anchor="{anchor}">{escape(content)}</text>'
        )

    def multiline_text(
        self,
        x: float,
        y: float,
        lines: list[str],
        size: float = 12,
        fill: str = "#1f2933",
        weight: str = "400",
        anchor: str = "start",
        line_height: float | None = None,
    ) -> None:
        if not lines:
            return
        line_height = line_height or size * 1.35
        self.add(
            f'<text x="{x:.2f}" y="{y:.2f}" font-family="Microsoft YaHei, SimHei, Arial, sans-serif" '
            f'font-size="{size}" font-weight="{weight}" fill="{fill}" text-anchor="{anchor}">'
        )
        for idx, line in enumerate(lines):
            dy = 0 if idx == 0 else line_height
            self.add(f'<tspan x="{x:.2f}" dy="{dy:.2f}">{escape(line)}</tspan>')
        self.add("</text>")

    def wrap_label(self, text: str, max_chars: int) -> list[str]:
        text = clean(text)
        if not text:
            return []
        chunks: list[str] = []
        current = ""
        for char in text:
            current += char
            if len(current) >= max_chars:
                chunks.append(current)
                current = ""
        if current:
            chunks.append(current)
        return chunks[:3]


def load_project(input_path: Path) -> tuple[dict[str, str], Topography, list[dict[str, str]], list[dict[str, str]]]:
    source = TableSource(input_path)
    section_rows = source.read("section_info")
    if not section_rows:
        raise FileNotFoundError(f"缺少 section_info 表：{input_path}")
    section = section_rows[0]
    section_id = clean(section.get("section_id"), "D")
    total_length = to_float(section.get("total_length_m"), None)

    topo_rows = [row for row in source.read("topography") if clean(row.get("section_id"), section_id) == section_id]
    topo_points = []
    for row in topo_rows:
        distance = to_float(row.get("distance_m"), None)
        elevation = to_float(row.get("elevation_m"), None)
        if distance is not None and elevation is not None:
            topo_points.append((distance, elevation))
    if total_length is None:
        total_length = max([d for d, _ in topo_points], default=100.0)
        section["total_length_m"] = str(total_length)
    topo = Topography(topo_points, total_length)

    units = [row for row in source.read("units") if clean(row.get("section_id"), section_id) == section_id]
    if not units:
        raise FileNotFoundError(f"缺少有效 units 分层数据：{input_path}")
    units.sort(key=lambda row: to_float(row.get("from_m"), 0.0) or 0.0)

    structures = [
        row for row in source.read("structures")
        if clean(row.get("section_id"), section_id) == section_id
    ]
    return section, topo, units, structures


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a geological section SVG from CSV tables.")
    parser.add_argument("--input", default="data/tables", help="CSV 表格目录，或包含同名工作表的 .xlsx 文件")
    parser.add_argument("--output", default="outputs/section_D.svg", help="输出 SVG 文件")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    input_dir = Path(args.input)
    output_file = Path(args.output)
    if not input_dir.is_absolute():
        input_dir = root / input_dir
    if not output_file.is_absolute():
        output_file = root / output_file

    section, topo, units, structures = load_project(input_dir)
    svg = SvgBuilder(section, topo, units, structures).render()
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text(svg, encoding="utf-8")
    print(f"Generated: {output_file}")


if __name__ == "__main__":
    main()
